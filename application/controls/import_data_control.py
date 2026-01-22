import uuid
import threading
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
from application.controls.auth_control import hash_password
from application.entities2 import *
from database.models import *
from database.base import get_session

ALL_IMPORT_JOBS = {
    "example": {
        "institution_id": 1,
        "import_users": {
            "total": 10,
            "success": 5,
            "failed": 5,
            "errors": [
                {
                    "row": 2,
                    "error": "Invalid email format"
                },
            ],
        },
        "import_venues": {},
        "import_courses": {},
        "assign_courses": {},
        "import_classes": {},
        "thread": None,
    }
}

def submit_import_data_job(institution_id: int, file_data: bytes) -> str:
    job_id = uuid.uuid4().hex
    thread = threading.Thread(target=process_excel_data, args=(job_id, file_data))
    thread.daemon = True
    ALL_IMPORT_JOBS[job_id] = {
        "institution_id": institution_id,
        "import_users": {"ws_name": "Import Users", "total": 0, "success": 0, "failed": 0, "errors": []},
        "import_venues": {"ws_name": "Import Venues", "total": 0, "success": 0, "failed": 0, "errors": []},
        "import_courses": {"ws_name": "Import Courses", "total": 0, "success": 0, "failed": 0, "errors": []},
        "assign_courses": {"ws_name": "Assign Courses", "total": 0, "success": 0, "failed": 0, "errors": []},
        "import_classes": {"ws_name": "Import Classes", "total": 0, "success": 0, "failed": 0, "errors": []},
    }
    thread.start()
    return job_id

def process_excel_data(job_id: str, file_data: bytes):
    job_state = ALL_IMPORT_JOBS[job_id]
    try:
        wb = load_workbook(BytesIO(file_data))
        import_tasks = ["import_users", "import_venues", "import_courses", "assign_courses", "import_classes"]

        # Update count first
        for task in import_tasks:
            ws_name = job_state[task]["ws_name"]
            if ws_name not in wb.sheetnames:
                job_state[task]["errors"].append({"row": 0, "error": f"Worksheet {ws_name} not found."})
                continue
            job_state[task]["total"] = wb[job_state[task]["ws_name"]].max_row - 1
        if any(job_state[task]["errors"] for task in import_tasks):
            return

        # Process each sheet into models
        def commit_to_db(task_name: str, enum_items):
            for row_num, item in enum_items:
                try:
                    with get_session() as session:
                        session.add(item)
                    job_state[task_name]["success"] += 1
                except Exception as e:
                    job_state[task_name]["failed"] += 1
                    job_state[task_name]["errors"].append({"row": row_num, "error": str(e.orig)})

        users = parse_user_sheet(job_id, wb["Import Users"])
        commit_to_db("import_users", users)

        venues = parse_venue_sheet(job_id, wb["Import Venues"])
        commit_to_db("import_venues", venues)

        courses = parse_course_sheet(job_id, wb["Import Courses"])
        commit_to_db("import_courses", courses)

        assignments = parse_assignment_sheet(job_id, wb["Assign Courses"])
        commit_to_db("assign_courses", assignments)

        classes = parse_class_sheet(job_id, wb["Import Classes"])
        commit_to_db("import_classes", classes)

    finally:
        time.sleep(30) # Give about 30s before erasing all records of the run
        ALL_IMPORT_JOBS.pop(job_id)

def parse_user_sheet(job_id: str, ws: Worksheet) -> List[User]:
    task_name = "import_users"
    # Columns are: Role, Name, Age, Gender, Email, Phone Number, Password
    base_info = {
        "institution_id": ALL_IMPORT_JOBS[job_id]["institution_id"],
    }
    headers = ['role', 'name', 'age', 'gender', 'email', 'phone_number', 'password']
    users = []
    for idx, row in enumerate(ws.iter_rows(), 1):
        if idx == 1:
            continue
        try:
            zipped_data = dict(zip(headers, [cell.value for cell in row]))
            zipped_data["password_hash"] = hash_password(zipped_data.pop("password"))
            users.append((idx, User(**base_info, **zipped_data)))
        except Exception as e:
            ALL_IMPORT_JOBS[job_id][task_name]["failed"] += 1
            ALL_IMPORT_JOBS[job_id][task_name]["errors"].append({"row": idx, "error": str(e)})
    return users

def parse_venue_sheet(job_id: str, ws: Worksheet) -> List[Venue]:
    task_name = "import_venues"
    base_info = {
        "institution_id": ALL_IMPORT_JOBS[job_id]["institution_id"],
    }
    headers = ['name', 'capacity']
    venues = []
    for idx, row in enumerate(ws.iter_rows(), 1):
        if idx == 1:
            continue
        try:
            zipped_data = dict(zip(headers, [cell.value for cell in row]))
            venues.append((idx, Venue(**base_info, **zipped_data)))
        except Exception as e:
            ALL_IMPORT_JOBS[job_id][task_name]["failed"] += 1
            ALL_IMPORT_JOBS[job_id][task_name]["errors"].append({"row": idx, "error": str(e)})
    return venues

def parse_course_sheet(job_id: str, ws: Worksheet) -> List[Course]:
    task_name = "import_courses"
    base_info = {
        "institution_id": ALL_IMPORT_JOBS[job_id]["institution_id"],
    }
    headers = ["code", "name", "description", "credits"]
    courses = []
    for idx, row in enumerate(ws.iter_rows(), 1):
        if idx == 1:
            continue
        try:
            zipped_data = dict(zip(headers, [cell.value for cell in row]))
            courses.append((idx, Course(**base_info, **zipped_data)))
        except Exception as e:
            ALL_IMPORT_JOBS[job_id][task_name]["failed"] += 1
            ALL_IMPORT_JOBS[job_id][task_name]["errors"].append({"row": idx, "error": str(e)})
    return courses

def parse_assignment_sheet(job_id: str, ws: Worksheet) -> List[CourseUser]:
    task_name = "assign_courses"
    inst_id = ALL_IMPORT_JOBS[job_id]["institution_id"]
    headers = ["email", "course_code", "semester_name"]
    assignments = []

    with get_session() as session:
        course_model = CourseModel(session)
        sem_model = SemesterModel(session)
        user_model = UserModel(session)
        courses = course_model.get_all(institution_id=inst_id)
        course_code_to_id = {course.code: course.course_id for course in courses}
        sems = sem_model.get_all(institution_id=inst_id)
        sem_name_to_id = {sem.name: sem.semester_id for sem in sems}
        users = user_model.get_all(institution_id=inst_id)
        user_email_to_id = {user.email: user.user_id for user in users}

    for idx, row in enumerate(ws.iter_rows(), 1):
        if idx == 1:
            continue
        try:
            zipped_data = dict(zip(headers, [cell.value for cell in row]))
            zipped_data["course_id"] = course_code_to_id.get(zipped_data.pop("course_code"), None)
            if zipped_data["course_id"] is None:
                raise ValueError("Invalid course code")
            zipped_data["semester_id"] = sem_name_to_id.get(zipped_data.pop("semester_name"), None)
            if zipped_data["semester_id"] is None:
                raise ValueError("Invalid semester name")
            zipped_data["user_id"] = user_email_to_id.get(zipped_data.pop("email"), None)
            if zipped_data["user_id"] is None:
                raise ValueError("Invalid user email")
            assignments.append((idx, CourseUser(**zipped_data)))
        except Exception as e:
            ALL_IMPORT_JOBS[job_id][task_name]["failed"] += 1
            ALL_IMPORT_JOBS[job_id][task_name]["errors"].append({"row": idx, "error": str(e)})
    return assignments

def parse_class_sheet(job_id: str, ws: Worksheet) -> List[Class]:
    task_name = "import_classes"
    inst_id = ALL_IMPORT_JOBS[job_id]["institution_id"]
    headers = ["course_code", "semester_name", "venue_name", "lecturer_email", "start_time", "end_time"]
    classes = []

    # Preload data for lookups
    with get_session() as session:
        course_model = CourseModel(session)
        sem_model = SemesterModel(session)
        venue_model = VenueModel(session)
        user_model = UserModel(session)
        courses = course_model.get_all(institution_id=inst_id)
        course_code_to_id = {course.code: course.course_id for course in courses}
        sems = sem_model.get_all(institution_id=inst_id)
        sem_name_to_id = {sem.name: sem.semester_id for sem in sems}
        venues = venue_model.get_all(institution_id=inst_id)
        venue_name_to_id = {venue.name: venue.venue_id for venue in venues}
        lecturers = user_model.get_all(institution_id=inst_id, role="lecturer")
        lecturer_email_to_id = {lecturer.email: lecturer.user_id for lecturer in lecturers}

    for idx, row in enumerate(ws.iter_rows(), 1):
        if idx == 1:
            continue
        try:
            zipped_data = dict(zip(headers, [cell.value for cell in row]))
            zipped_data["course_id"] = course_code_to_id.get(zipped_data.pop("course_code"), None)
            if zipped_data["course_id"] is None:
                raise ValueError("Invalid course code")
            zipped_data["semester_id"] = sem_name_to_id.get(zipped_data.pop("semester_name"), None)
            if zipped_data["semester_id"] is None:
                raise ValueError("Invalid semester name")
            zipped_data["venue_id"] = venue_name_to_id.get(zipped_data.pop("venue_name"), None)
            if zipped_data["venue_id"] is None:
                raise ValueError("Invalid venue name")
            zipped_data["lecturer_id"] = lecturer_email_to_id.get(zipped_data.pop("lecturer_email"), None)
            if zipped_data["lecturer_id"] is None:
                raise ValueError("Invalid lecturer email")
            classes.append((idx, Class(**zipped_data)))
        except Exception as e:
            ALL_IMPORT_JOBS[job_id][task_name]["failed"] += 1
            ALL_IMPORT_JOBS[job_id][task_name]["errors"].append({"row": idx, "error": str(e)})
    return classes

if __name__ == "__main__":
    ALL_IMPORT_JOBS[1] = {
        "institution_id": 2,
        "import_users": {"ws_name": "Import Users", "total": 0, "success": 0, "failed": 0, "errors": []},
        "import_venues": {"ws_name": "Import Venues", "total": 0, "success": 0, "failed": 0, "errors": []},
        "import_courses": {"ws_name": "Import Courses", "total": 0, "success": 0, "failed": 0, "errors": []},
        "assign_courses": {"ws_name": "Assign Courses", "total": 0, "success": 0, "failed": 0, "errors": []},
        "import_classes": {"ws_name": "Import Classes", "total": 0, "success": 0, "failed": 0, "errors": []},
    }
    with open("static/import-data.xlsx", "rb") as f:
        bin_data = f.read()
    process_excel_data(1, bin_data)
